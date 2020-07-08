# -*- coding:utf-8 -*-
# @time   :2020/6/30  11:23
# @Author :sunzhe
# @Email  :15967579213@163.com
# @File   :run.py
#读取文件
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="gb18030")

from openpyxl import load_workbook
from class_automations.class_homework_R_W_excel import read_data
from class_automations.http_Request_case import http_request
all_case = read_data('login_test_1.xlsx', 'recharge')
login_data=all_case[0]
ip = 'http://120.78.128.25:8766'
log_response = http_request(ip+login_data[4],eval(login_data[5]))
mumber_id = log_response['data']['id']
# print(type(login_data[0][6]))
wb = load_workbook('login_test_1.xlsx')
ws = wb['recharge']
if log_response['code'] == eval(login_data[6])['code'] and log_response['msg'] == eval(login_data[6])['msg']:
    ws.cell(row=2, column=8).value = str(log_response)
    ws.cell(row=2, column=9).value = 'pass'
else:
    ws.cell(row=2, column=8).value = str(log_response)
    ws.cell(row=2, column=9).value = 'false'
wb.save('login_test_1.xlsx')
#执行测试用例
for index in range(1,len(all_case)):
    test_data = all_case[index]
    token = log_response['data']['token_info']['token']
    response = http_request(ip+test_data[4],eval(test_data[5]),'Bearer ' + token)
    print(response)
    if response['code'] == eval(test_data[6])['code'] and response['msg'] == eval(test_data[6])['msg']:
        ws.cell(row=index+2, column=8).value = str(response)
        ws.cell(row=index+2, column=9).value = 'pass'
    else:
        ws.cell(row=index+2, column=8).value = str(response)
        ws.cell(row=index+2, column=9).value = 'false'
    wb.save('login_test_1.xlsx')


    # def run(file_name, list_name, column_1, column_2):
    #     global Token  # 声明函数外的token和函数内的token是一个值
    #     all_case = read_data(file_name, list_name)  # 此时 我们就调用到了所有的测试用例数据，内嵌在表格里面
    #     for test_data in all_case:  # 在http_requests进行请求的时候#每一行，所有的测试用例
    #         ip = 'http://120.78.128.25:8766'  # 可能会随着环境的变化会跟着变化
    #         res = http_request(ip + test_data[4], eval(test_data[5]), Token, test_data[3])  # 获取第一行测试数据中的索引值4 url和5参数
    #         # if text_data[0]==1:   #判断是不是登录这条用例,根据登录这条用例中索引值对应的数据
    #         # 或者if text_data[1]=='登录'#   ==是比较运算符，判读两边是否相等
    #         if 'login' in test_data[4]:  # in成员运算符#此时此刻我的Token = None
    #             Token = 'Bearer ' + res['data']['token_info']['token']  # 获取登录后的token\
    #         # print('此时执行的结果为：',format(res))
    #         # wb = load_workbook(file_name)
    #         # sheet = wb[list_name]
    #         # sheet.cell(row=test_data[0]+1,column=column_1).value = str(res) #此时的test_data[0]=1,2,3.....
    #         # actual = {'code':res['code'],'msg':res['msg']}
    #         #
    #         write_data(file_name, list_name, test_data[0] + 1, column_1, str(res))
    #         actual = {'code': res['code'], 'msg': res['msg']}
    #         if eval(test_data[6]) == actual:
    #             print('测试用例通过')
    #             write_data(file_name, list_name, test_data[0] + 1, column_2, 'PASS')
    #         else:
    #             print('测试用例不通过')
    #             write_data(file_name, list_name, test_data[0] + 1, column_2, 'FAIL')
    #         # wb.save(file_name)  # 一定要记得的保存
    #
    #
    # run('text_002.xlsx', 'recharge', 8, 9)
