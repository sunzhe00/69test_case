# -*- coding:utf-8 -*-
# @time   :2020/6/21  11:53
# @Author :sunzhe
# @Email  :15967579213@163.com
# @File   :class_homework_R_W_excel.py

from openpyxl import load_workbook
def read_data(file_name,list_name):
    wb = load_workbook(file_name)
    ws = wb[list_name]
    data = []   #存储所有行的数据
    for rows in range(2,ws.max_row+1):#从第二行开始读取数据，到最大行+1
        list_1 = []    #存储每次循环行的数据
        for col in range(1,ws.max_column+1):#从第一列开始读取数据，到最大列+1
            values = ws.cell(row=rows,column=col).value
            list_1.append(values)
        data.append(list_1)
    return data
if __name__ == '__main__':
    read_data('login_test_1.xlsx', 'recharge')
    print(read_data('login_test_1.xlsx', 'recharge'))
